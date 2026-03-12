"""
Servicio para generación de reportes en Excel.
Usa openpyxl para generar archivos xlsx.
"""
import io
from datetime import date
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generar_excel_reporte(datos: dict) -> Optional[bytes]:
    """
    Genera un Excel del reporte de proyecto.

    Args:
        datos: Diccionario con los datos del reporte (de reporte_service.obtener_datos_reporte)

    Returns:
        bytes del Excel generado, o None si openpyxl no está disponible
    """
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    subheader_font = Font(bold=True)
    subheader_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    currency_format = '$#,##0.00'
    percent_format = '0%'

    # === Hoja 1: Resumen ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    proyecto = datos.get('proyecto', {})
    periodo = datos.get('periodo', {})
    resumen = datos.get('resumen', {})

    # Titulo
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = f"Reporte: {proyecto.get('nombre', 'Proyecto')}"
    ws_resumen['A1'].font = Font(bold=True, size=16)

    ws_resumen['A2'] = f"Periodo: {periodo.get('desde', '')} - {periodo.get('hasta', '')}"
    ws_resumen['A3'] = f"Ubicacion: {proyecto.get('ubicacion', '-')}"

    # Resumen general
    ws_resumen['A5'] = "RESUMEN GENERAL"
    ws_resumen['A5'].font = subheader_font

    resumen_data = [
        ("Avance General", f"{proyecto.get('porcentaje_avance', 0):.0f}%"),
        ("Estado", proyecto.get('estado', '-')),
        ("Etapas Totales", resumen.get('total_etapas', 0)),
        ("Etapas Completadas", resumen.get('etapas_completadas', 0)),
        ("Total Materiales", resumen.get('total_materiales_asignados', 0)),
        ("Costo Materiales", resumen.get('costo_materiales', 0)),
        ("Total Gastos", resumen.get('total_gastos', 0)),
        ("Costo Total", resumen.get('costo_materiales', 0) + resumen.get('total_gastos', 0)),
    ]

    row = 6
    for label, value in resumen_data:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'B{row}'] = value
        if 'Costo' in label or 'Gasto' in label:
            ws_resumen[f'B{row}'].number_format = currency_format
        row += 1

    # Cliente y Supervisor
    row += 1
    ws_resumen[f'A{row}'] = "INFORMACION DE CONTACTO"
    ws_resumen[f'A{row}'].font = subheader_font
    row += 1

    if proyecto.get('cliente'):
        ws_resumen[f'A{row}'] = "Cliente"
        ws_resumen[f'B{row}'] = proyecto['cliente'].get('nombre', '-')
        row += 1
        ws_resumen[f'A{row}'] = "Email"
        ws_resumen[f'B{row}'] = proyecto['cliente'].get('email', '-')
        row += 1

    if proyecto.get('supervisor'):
        ws_resumen[f'A{row}'] = "Supervisor"
        ws_resumen[f'B{row}'] = proyecto['supervisor'].get('nombre', '-')
        row += 1

    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 30

    # === Hoja 2: Etapas ===
    ws_etapas = wb.create_sheet("Etapas")
    etapas = datos.get('etapas', [])

    # Headers
    headers = ["Etapa", "Estado", "Avance %", "Items", "Fecha Inicio Est.", "Fecha Fin Est."]
    for col, header in enumerate(headers, 1):
        cell = ws_etapas.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row_idx, etapa in enumerate(etapas, 2):
        ws_etapas.cell(row=row_idx, column=1, value=etapa.get('nombre', '')).border = border
        ws_etapas.cell(row=row_idx, column=2, value=etapa.get('estado', '')).border = border
        cell_avance = ws_etapas.cell(row=row_idx, column=3, value=etapa.get('porcentaje_avance', 0) / 100)
        cell_avance.number_format = percent_format
        cell_avance.border = border
        ws_etapas.cell(row=row_idx, column=4, value=etapa.get('items_total', 0)).border = border
        ws_etapas.cell(row=row_idx, column=5, value=etapa.get('fecha_inicio_est', '')).border = border
        ws_etapas.cell(row=row_idx, column=6, value=etapa.get('fecha_fin_est', '')).border = border

    # Ajustar anchos
    for col in range(1, 7):
        ws_etapas.column_dimensions[get_column_letter(col)].width = 18

    # === Hoja 3: Items de Trabajo ===
    ws_items = wb.create_sheet("Items de Trabajo")

    headers = ["Etapa", "Item", "Unidad", "Cant. Estimada", "Cant. Ejecutada", "Avance %", "Precio Unit."]
    for col, header in enumerate(headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row_idx = 2
    for etapa in etapas:
        for item in etapa.get('items', []):
            ws_items.cell(row=row_idx, column=1, value=etapa.get('nombre', '')).border = border
            ws_items.cell(row=row_idx, column=2, value=item.get('nombre', '')).border = border
            ws_items.cell(row=row_idx, column=3, value=item.get('unidad', '')).border = border
            ws_items.cell(row=row_idx, column=4, value=item.get('cantidad_estimada', 0)).border = border
            ws_items.cell(row=row_idx, column=5, value=item.get('cantidad_ejecutada', 0)).border = border
            cell_avance = ws_items.cell(row=row_idx, column=6, value=item.get('porcentaje', 0) / 100)
            cell_avance.number_format = percent_format
            cell_avance.border = border
            cell_precio = ws_items.cell(row=row_idx, column=7, value=item.get('precio_unitario', 0))
            cell_precio.number_format = currency_format
            cell_precio.border = border
            row_idx += 1

    for col in range(1, 8):
        ws_items.column_dimensions[get_column_letter(col)].width = 15

    # === Hoja 4: Materiales ===
    ws_materiales = wb.create_sheet("Materiales")
    materiales = datos.get('materiales', [])

    headers = ["Material", "Codigo", "Cantidad", "Unidad", "Etapa", "Fecha", "Costo Est."]
    for col, header in enumerate(headers, 1):
        cell = ws_materiales.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, mat in enumerate(materiales, 2):
        ws_materiales.cell(row=row_idx, column=1, value=mat.get('nombre', '')).border = border
        ws_materiales.cell(row=row_idx, column=2, value=mat.get('codigo', '')).border = border
        ws_materiales.cell(row=row_idx, column=3, value=mat.get('cantidad', 0)).border = border
        ws_materiales.cell(row=row_idx, column=4, value=mat.get('unidad', '')).border = border
        ws_materiales.cell(row=row_idx, column=5, value=mat.get('etapa', '')).border = border
        ws_materiales.cell(row=row_idx, column=6, value=mat.get('fecha', '')).border = border
        cell_costo = ws_materiales.cell(row=row_idx, column=7, value=mat.get('costo_estimado', 0))
        cell_costo.number_format = currency_format
        cell_costo.border = border

    # Total
    if materiales:
        row_idx = len(materiales) + 2
        ws_materiales.cell(row=row_idx, column=6, value="TOTAL").font = subheader_font
        cell_total = ws_materiales.cell(row=row_idx, column=7, value=resumen.get('costo_materiales', 0))
        cell_total.number_format = currency_format
        cell_total.font = subheader_font

    for col in range(1, 8):
        ws_materiales.column_dimensions[get_column_letter(col)].width = 15

    # === Hoja 5: Gastos ===
    ws_gastos = wb.create_sheet("Gastos")
    gastos = datos.get('gastos', [])

    headers = ["Fecha", "Descripcion", "Categoria", "Monto", "Comprobante"]
    for col, header in enumerate(headers, 1):
        cell = ws_gastos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, gasto in enumerate(gastos, 2):
        ws_gastos.cell(row=row_idx, column=1, value=gasto.get('fecha', '')).border = border
        ws_gastos.cell(row=row_idx, column=2, value=gasto.get('descripcion', '')).border = border
        ws_gastos.cell(row=row_idx, column=3, value=gasto.get('categoria', '')).border = border
        cell_monto = ws_gastos.cell(row=row_idx, column=4, value=gasto.get('monto', 0))
        cell_monto.number_format = currency_format
        cell_monto.border = border
        ws_gastos.cell(row=row_idx, column=5, value=gasto.get('numero_comprobante', '')).border = border

    # Total
    if gastos:
        row_idx = len(gastos) + 2
        ws_gastos.cell(row=row_idx, column=3, value="TOTAL").font = subheader_font
        cell_total = ws_gastos.cell(row=row_idx, column=4, value=resumen.get('total_gastos', 0))
        cell_total.number_format = currency_format
        cell_total.font = subheader_font

    for col in range(1, 6):
        ws_gastos.column_dimensions[get_column_letter(col)].width = 18

    # === Hoja 6: Equipos ===
    ws_equipos = wb.create_sheet("Equipos")
    equipos = datos.get('equipos', [])

    headers = ["Equipo", "Codigo", "Tipo", "Marca", "Modelo", "F. Asignacion", "F. Devolucion"]
    for col, header in enumerate(headers, 1):
        cell = ws_equipos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, equipo in enumerate(equipos, 2):
        ws_equipos.cell(row=row_idx, column=1, value=equipo.get('nombre', '')).border = border
        ws_equipos.cell(row=row_idx, column=2, value=equipo.get('codigo', '')).border = border
        ws_equipos.cell(row=row_idx, column=3, value=equipo.get('tipo', '')).border = border
        ws_equipos.cell(row=row_idx, column=4, value=equipo.get('marca', '')).border = border
        ws_equipos.cell(row=row_idx, column=5, value=equipo.get('modelo', '')).border = border
        ws_equipos.cell(row=row_idx, column=6, value=equipo.get('fecha_asignacion', '')).border = border
        ws_equipos.cell(row=row_idx, column=7, value=equipo.get('fecha_devolucion_real', '')).border = border

    for col in range(1, 8):
        ws_equipos.column_dimensions[get_column_letter(col)].width = 15

    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()
